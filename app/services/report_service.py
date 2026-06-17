from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime, timedelta, timezone
from collections import defaultdict
import os
import uuid
from decimal import Decimal
import numpy as np
import pandas as pd
from sqlalchemy import select, and_, func, case, extract
from sqlalchemy.ext.asyncio import AsyncSession
from ..database import async_session_maker, sync_engine
from ..models.anomaly import Anomaly
from ..models.ticket import WorkOrder
from ..models.playbook import PlaybookExecution
from ..models.user import Team
from ..models.report import DailyReport
from ..config import settings
from ..utils.logger import logger


class ReportService:
    def __init__(self):
        self._export_dir = settings.REPORT_EXPORT_DIR

    async def generate_daily_report(
        self, report_date: Optional[datetime] = None, save: bool = True
    ) -> DailyReport:
        if report_date is None:
            report_date = datetime.now(timezone.utc).replace(
                hour=0, minute=0, second=0, microsecond=0
            )

        report_start = report_date
        report_end = report_start + timedelta(days=1)

        logger.info(f"Generating daily report for {report_start.date()}")

        async with async_session_maker() as db:
            anomaly_stats = await self._get_anomaly_stats(db, report_start, report_end)
            wo_stats = await self._get_work_order_stats(db, report_start, report_end)
            resolution_stats = await self._get_resolution_stats(db, report_start, report_end)
            top_stats = await self._get_top_stats(db, report_start, report_end)
            trend_data = await self._get_trend_data(db, report_start, report_end)
            hourly_dist = await self._get_hourly_distribution(db, report_start, report_end)
            impact_breakdown = await self._get_impact_scope_breakdown(db, report_start, report_end)
            system_summary = await self._get_system_summary(db, report_start, report_end)
            team_summary = await self._get_team_summary(db, report_start, report_end)
            playbook_summary = await self._get_playbook_summary(db, report_start, report_end)
            notable_events = await self._get_notable_events(db, report_start, report_end)

            repeat_rate = await self._calculate_repeat_rate(db, anomaly_stats, report_start, report_end)

            sla_compliance = self._calculate_sla_compliance(wo_stats, resolution_stats)

            recommendations = self._generate_recommendations(
                anomaly_stats, resolution_stats, top_stats, sla_compliance
            )

        report = DailyReport(
            report_date=report_start,
            **anomaly_stats,
            **wo_stats,
            **resolution_stats,
            repeat_rate=Decimal(str(repeat_rate)),
            top_anomaly_types=top_stats["anomaly_types"],
            top_affected_systems=top_stats["systems"],
            top_root_causes=top_stats["root_causes"],
            sla_compliance_rate=Decimal(str(sla_compliance)),
            impact_scope_breakdown=impact_breakdown,
            trend_data=trend_data,
            anomaly_hourly_distribution=hourly_dist,
            system_summary=system_summary,
            team_summary=team_summary,
            playbook_execution_summary=playbook_summary,
            notable_events=notable_events,
            recommendations=recommendations,
            generated_by="auto",
        )

        if save:
            async with async_session_maker() as db:
                existing = await db.execute(
                    select(DailyReport).where(DailyReport.report_date == report_start)
                )
                existing_report = existing.scalar_one_or_none()
                if existing_report:
                    for key, value in report.__dict__.items():
                        if not key.startswith("_") and key != "id":
                            setattr(existing_report, key, value)
                    await db.commit()
                    await db.refresh(existing_report)
                    report = existing_report
                else:
                    db.add(report)
                    await db.commit()
                    await db.refresh(report)

        logger.info(f"Daily report generated for {report_start.date()}: {report.total_anomalies} anomalies")
        return report

    async def _get_anomaly_stats(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> Dict[str, int]:
        stats = {
            "total_anomalies": 0,
            "critical_anomalies": 0,
            "high_anomalies": 0,
            "medium_anomalies": 0,
            "low_anomalies": 0,
            "new_anomalies": 0,
            "resolved_anomalies": 0,
            "unresolved_anomalies": 0,
            "auto_resolved": 0,
            "manual_resolved": 0,
            "playbook_resolved": 0,
            "escalated_count": 0,
        }

        total_result = await db.execute(
            select(
                func.count(Anomaly.id),
                func.sum(case((Anomaly.severity == "critical", 1), else_=0)),
                func.sum(case((Anomaly.severity == "high", 1), else_=0)),
                func.sum(case((Anomaly.severity == "medium", 1), else_=0)),
                func.sum(case((Anomaly.severity == "low", 1), else_=0)),
                func.sum(case((Anomaly.status.in_(["resolved", "closed"]), 1), else_=0)),
            ).where(
                and_(Anomaly.detected_time >= start, Anomaly.detected_time < end)
            )
        )
        row = total_result.fetchone()
        if row:
            stats["total_anomalies"] = int(row[0] or 0)
            stats["critical_anomalies"] = int(row[1] or 0)
            stats["high_anomalies"] = int(row[2] or 0)
            stats["medium_anomalies"] = int(row[3] or 0)
            stats["low_anomalies"] = int(row[4] or 0)
            stats["resolved_anomalies"] = int(row[5] or 0)
            stats["unresolved_anomalies"] = stats["total_anomalies"] - stats["resolved_anomalies"]

        new_result = await db.execute(
            select(func.count(Anomaly.id)).where(
                and_(
                    Anomaly.status == "open",
                    Anomaly.detected_time >= start,
                    Anomaly.detected_time < end,
                )
            )
        )
        stats["new_anomalies"] = int(new_result.scalar() or 0)

        resolved_detail = await db.execute(
            select(
                func.sum(case((Anomaly.resolution_method == "auto", 1), else_=0)),
                func.sum(case((Anomaly.resolution_method == "manual", 1), else_=0)),
                func.sum(case((Anomaly.resolution_method == "playbook", 1), else_=0)),
            ).where(
                and_(
                    Anomaly.resolved_time >= start,
                    Anomaly.resolved_time < end,
                    Anomaly.status.in_(["resolved", "closed"]),
                )
            )
        )
        r_row = resolved_detail.fetchone()
        if r_row:
            stats["auto_resolved"] = int(r_row[0] or 0)
            stats["manual_resolved"] = int(r_row[1] or 0)
            stats["playbook_resolved"] = int(r_row[2] or 0)

        escalated_result = await db.execute(
            select(func.count(WorkOrder.id)).where(
                and_(
                    WorkOrder.created_at >= start,
                    WorkOrder.created_at < end,
                    WorkOrder.is_escalated == True,
                )
            )
        )
        stats["escalated_count"] = int(escalated_result.scalar() or 0)

        return stats

    async def _get_work_order_stats(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> Dict[str, int]:
        stats = {"total_work_orders": 0, "completed_work_orders": 0, "sla_breach_count": 0}

        result = await db.execute(
            select(
                func.count(WorkOrder.id),
                func.sum(case((WorkOrder.status.in_(["completed", "closed"]), 1), else_=0)),
            ).where(
                and_(WorkOrder.created_at >= start, WorkOrder.created_at < end)
            )
        )
        row = result.fetchone()
        if row:
            stats["total_work_orders"] = int(row[0] or 0)
            stats["completed_work_orders"] = int(row[1] or 0)

        breach_result = await db.execute(
            select(func.count(WorkOrder.id)).where(
                and_(
                    WorkOrder.created_at >= start,
                    WorkOrder.created_at < end,
                    WorkOrder.sla_deadline.isnot(None),
                    WorkOrder.resolved_at.isnot(None),
                    WorkOrder.resolved_at > WorkOrder.sla_deadline,
                )
            )
        )
        stats["sla_breach_count"] = int(breach_result.scalar() or 0)

        return stats

    async def _get_resolution_stats(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> Dict[str, Any]:
        result = await db.execute(
            select(WorkOrder.actual_resolution_minutes).where(
                and_(
                    WorkOrder.resolved_at >= start,
                    WorkOrder.resolved_at < end,
                    WorkOrder.actual_resolution_minutes.isnot(None),
                )
            )
        )
        minutes = [row[0] for row in result.fetchall() if row[0]]

        if not minutes:
            return {
                "avg_resolution_minutes": None,
                "median_resolution_minutes": None,
                "p95_resolution_minutes": None,
            }

        arr = np.array(minutes, dtype=float)
        return {
            "avg_resolution_minutes": Decimal(str(round(float(np.mean(arr)), 2))),
            "median_resolution_minutes": Decimal(str(round(float(np.median(arr)), 2))),
            "p95_resolution_minutes": Decimal(str(round(float(np.percentile(arr, 95)), 2))),
        }

    async def _get_top_stats(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> Dict[str, List[Dict]]:
        anomaly_types = []
        at_result = await db.execute(
            select(Anomaly.anomaly_type, func.count(Anomaly.id)).where(
                and_(Anomaly.detected_time >= start, Anomaly.detected_time < end)
            ).group_by(Anomaly.anomaly_type).order_by(func.count(Anomaly.id).desc()).limit(10)
        )
        for type_name, count in at_result.fetchall():
            anomaly_types.append({"type": type_name, "count": int(count)})

        systems = []
        sys_result = await db.execute(
            select(Anomaly.system_name, func.count(Anomaly.id)).where(
                and_(Anomaly.detected_time >= start, Anomaly.detected_time < end)
            ).group_by(Anomaly.system_name).order_by(func.count(Anomaly.id).desc()).limit(10)
        )
        for name, count in sys_result.fetchall():
            systems.append({"system": name, "count": int(count)})

        root_causes = []
        rc_result = await db.execute(
            select(WorkOrder.root_cause_category, func.count(WorkOrder.id)).where(
                and_(
                    WorkOrder.created_at >= start,
                    WorkOrder.created_at < end,
                    WorkOrder.root_cause_category.isnot(None),
                )
            ).group_by(WorkOrder.root_cause_category).order_by(func.count(WorkOrder.id).desc()).limit(10)
        )
        for category, count in rc_result.fetchall():
            root_causes.append({"category": category, "count": int(count)})

        return {"anomaly_types": anomaly_types, "systems": systems, "root_causes": root_causes}

    async def _get_trend_data(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> List[Dict[str, Any]]:
        trend = []
        days = 7

        for i in range(days):
            day_start = start - timedelta(days=days - 1 - i)
            day_end = day_start + timedelta(days=1)

            result = await db.execute(
                select(func.count(Anomaly.id)).where(
                    and_(Anomaly.detected_time >= day_start, Anomaly.detected_time < day_end)
                )
            )
            count = int(result.scalar() or 0)
            trend.append({
                "date": day_start.date().isoformat(),
                "anomaly_count": count,
            })

        return trend

    async def _get_hourly_distribution(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> Dict[str, int]:
        result = await db.execute(
            select(
                extract("hour", Anomaly.detected_time),
                func.count(Anomaly.id),
            ).where(
                and_(Anomaly.detected_time >= start, Anomaly.detected_time < end)
            ).group_by(extract("hour", Anomaly.detected_time)).order_by(extract("hour", Anomaly.detected_time))
        )

        dist = {str(h): 0 for h in range(24)}
        for hour, count in result.fetchall():
            dist[str(int(hour))] = int(count)
        return dist

    async def _get_impact_scope_breakdown(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> List[Dict[str, Any]]:
        result = await db.execute(
            select(Anomaly.impact_scope, func.count(Anomaly.id)).where(
                and_(Anomaly.detected_time >= start, Anomaly.detected_time < end)
            ).group_by(Anomaly.impact_scope)
        )
        scope_labels = {
            "single": "单服务",
            "module": "模块级",
            "system": "系统级",
            "multi_system": "跨系统",
        }
        breakdown = []
        for scope, count in result.fetchall():
            breakdown.append({
                "scope": scope,
                "label": scope_labels.get(scope, scope),
                "count": int(count),
            })
        return breakdown

    async def _get_system_summary(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> List[Dict[str, Any]]:
        result = await db.execute(
            select(
                Anomaly.system_name,
                func.count(Anomaly.id),
                func.sum(case((Anomaly.severity == "critical", 1), else_=0)),
                func.sum(case((Anomaly.status.in_(["resolved", "closed"]), 1), else_=0)),
            ).where(
                and_(Anomaly.detected_time >= start, Anomaly.detected_time < end)
            ).group_by(Anomaly.system_name).order_by(func.count(Anomaly.id).desc()).limit(20)
        )

        summary = []
        for sys_name, total, critical, resolved in result.fetchall():
            total_i = int(total or 0)
            summary.append({
                "system": sys_name,
                "total": total_i,
                "critical": int(critical or 0),
                "resolved": int(resolved or 0),
                "resolution_rate": round(int(resolved or 0) / total_i * 100, 2) if total_i > 0 else 0,
            })
        return summary

    async def _get_team_summary(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> List[Dict[str, Any]]:
        result = await db.execute(
            select(
                Team.name,
                func.count(WorkOrder.id),
                func.sum(case((WorkOrder.status.in_(["completed", "closed"]), 1), else_=0)),
                func.sum(case((WorkOrder.is_escalated == True, 1), else_=0)),
            ).outerjoin(
                WorkOrder,
                and_(
                    WorkOrder.assigned_team_id == Team.id,
                    WorkOrder.created_at >= start,
                    WorkOrder.created_at < end,
                )
            ).group_by(Team.name).order_by(func.count(WorkOrder.id).desc())
        )

        summary = []
        for team_name, total, completed, escalated in result.fetchall():
            total_i = int(total or 0)
            summary.append({
                "team": team_name,
                "total": total_i,
                "completed": int(completed or 0),
                "escalated": int(escalated or 0),
                "completion_rate": round(int(completed or 0) / total_i * 100, 2) if total_i > 0 else 0,
            })
        return summary

    async def _get_playbook_summary(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> Dict[str, Any]:
        result = await db.execute(
            select(
                func.count(PlaybookExecution.id),
                func.sum(case((PlaybookExecution.status == "success", 1), else_=0)),
                func.sum(case((PlaybookExecution.status == "failed", 1), else_=0)),
                func.sum(case((PlaybookExecution.status == "rolled_back", 1), else_=0)),
                func.avg(PlaybookExecution.duration_seconds),
            ).where(
                and_(
                    PlaybookExecution.created_at >= start,
                    PlaybookExecution.created_at < end,
                )
            )
        )
        row = result.fetchone()
        total = int(row[0] or 0) if row else 0
        success = int(row[1] or 0) if row else 0
        return {
            "total_executions": total,
            "success_count": success,
            "failed_count": int(row[2] or 0) if row else 0,
            "rolled_back_count": int(row[3] or 0) if row else 0,
            "avg_duration_seconds": round(float(row[4]), 2) if row and row[4] else 0,
            "success_rate": round(success / total * 100, 2) if total > 0 else 0,
        }

    async def _get_notable_events(
        self, db: AsyncSession, start: datetime, end: datetime
    ) -> List[Dict[str, Any]]:
        result = await db.execute(
            select(Anomaly).where(
                and_(
                    Anomaly.detected_time >= start,
                    Anomaly.detected_time < end,
                    Anomaly.severity.in_(["critical", "high"]),
                )
            ).order_by(Anomaly.impact_score.desc()).limit(10)
        )
        events = []
        for anomaly in result.scalars().all():
            events.append({
                "anomaly_code": anomaly.anomaly_code,
                "title": anomaly.title,
                "system": anomaly.system_name,
                "severity": anomaly.severity,
                "impact_score": float(anomaly.impact_score) if anomaly.impact_score else 0,
                "status": anomaly.status,
                "detected_time": anomaly.detected_time.isoformat(),
            })
        return events

    async def _calculate_repeat_rate(
        self,
        db: AsyncSession,
        anomaly_stats: Dict[str, Any],
        start: datetime,
        end: datetime,
    ) -> float:
        total = anomaly_stats.get("total_anomalies", 0)
        if total == 0:
            return 0.0

        prev_start = start - timedelta(days=7)
        prev_end = start

        prev_result = await db.execute(
            select(Anomaly.system_name, Anomaly.anomaly_type).where(
                and_(Anomaly.detected_time >= prev_start, Anomaly.detected_time < prev_end)
            )
        )
        prev_pairs = set()
        for s, t in prev_result.fetchall():
            prev_pairs.add((s, t))

        curr_result = await db.execute(
            select(Anomaly.system_name, Anomaly.anomaly_type).where(
                and_(Anomaly.detected_time >= start, Anomaly.detected_time < end)
            )
        )
        repeated = 0
        for s, t in curr_result.fetchall():
            if (s, t) in prev_pairs:
                repeated += 1

        return round(repeated / total * 100, 2)

    def _calculate_sla_compliance(
        self, wo_stats: Dict[str, Any], resolution_stats: Dict[str, Any]
    ) -> float:
        total = wo_stats.get("total_work_orders", 0)
        breaches = wo_stats.get("sla_breach_count", 0)
        if total == 0:
            return 100.0
        return round((total - breaches) / total * 100, 2)

    def _generate_recommendations(
        self,
        anomaly_stats: Dict[str, Any],
        resolution_stats: Dict[str, Any],
        top_stats: Dict[str, Any],
        sla_compliance: float,
    ) -> List[str]:
        recs = []

        critical_ratio = anomaly_stats.get("critical_anomalies", 0) / max(
            anomaly_stats.get("total_anomalies", 1), 1
        )
        if critical_ratio > 0.1:
            recs.append(
                f"严重异常占比过高({critical_ratio*100:.1f}%)，建议排查Top受影响系统的稳定性问题。"
            )

        if sla_compliance < 95:
            recs.append(
                f"SLA达标率为{sla_compliance:.1f}%，低于95%的目标值，建议加强工单响应速度。"
            )

        avg_mins = resolution_stats.get("avg_resolution_minutes")
        if avg_mins and float(avg_mins) > 120:
            recs.append(
                f"平均修复时长{float(avg_mins):.0f}分钟，建议优化预案库并提升自动化覆盖率。"
            )

        repeat_rate = anomaly_stats.get("repeat_rate", 0)
        if isinstance(repeat_rate, Decimal):
            repeat_rate = float(repeat_rate)
        if repeat_rate and repeat_rate > 30:
            recs.append(
                f"异常重复率为{repeat_rate:.1f}%，建议针对重复异常制定专项治理方案。"
            )

        if top_stats.get("systems"):
            top_system = top_stats["systems"][0]
            if top_system["count"] > 20:
                recs.append(
                    f"系统{top_system['system']}当日异常数最高({top_system['count']}个)，建议重点关注。"
                )

        if not recs:
            recs.append("整体运营情况良好，继续保持当前运维策略。")

        return recs

    async def export_report_pdf(self, report_id: str) -> Optional[str]:
        async with async_session_maker() as db:
            result = await db.execute(
                select(DailyReport).where(DailyReport.id == report_id)
            )
            report = result.scalar_one_or_none()
            if report is None:
                return None

        return await self._generate_pdf_report(report)

    async def _generate_pdf_report(self, report: DailyReport) -> str:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, Image
        )
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib import font_manager
        import io

        try:
            pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
            font_name = "STSong-Light"
        except Exception:
            font_name = "Helvetica"

        date_str = report.report_date.strftime("%Y-%m-%d")
        filename = f"daily_report_{date_str}_{uuid.uuid4().hex[:8]}.pdf"
        filepath = os.path.join(self._export_dir, filename)

        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            leftMargin=2 * cm,
            rightMargin=2 * cm,
            topMargin=2.5 * cm,
            bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Title"],
            fontName=font_name,
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#1a5490"),
        )
        h2_style = ParagraphStyle(
            "CustomH2",
            parent=styles["Heading2"],
            fontName=font_name,
            fontSize=14,
            leading=18,
            textColor=colors.HexColor("#2c3e50"),
        )
        normal_style = ParagraphStyle(
            "CustomNormal",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=10,
            leading=14,
        )

        story = []

        story.append(Paragraph("运维异常检测系统 - 每日汇总报告", title_style))
        story.append(Spacer(1, 5 * mm))
        story.append(Paragraph(f"报告日期: {date_str}", normal_style))
        story.append(Paragraph(f"生成时间: {report.generated_at.strftime('%Y-%m-%d %H:%M:%S') if report.generated_at else ''}", normal_style))
        story.append(Spacer(1, 10 * mm))

        story.append(Paragraph("一、核心指标概览", h2_style))
        story.append(Spacer(1, 5 * mm))

        summary_data = [
            ["指标", "数值", "指标", "数值"],
            ["总异常数", str(report.total_anomalies), "总工单数", str(report.total_work_orders)],
            ["严重异常", str(report.critical_anomalies), "高危异常", str(report.high_anomalies)],
            ["已解决异常", str(report.resolved_anomalies), "解决率", f"{(report.resolved_anomalies / max(report.total_anomalies,1) * 100):.1f}%"],
            ["平均修复时长", f"{report.avg_resolution_minutes}分钟" if report.avg_resolution_minutes else "-",
             "P95修复时长", f"{report.p95_resolution_minutes}分钟" if report.p95_resolution_minutes else "-"],
            ["重复率", f"{report.repeat_rate}%" if report.repeat_rate else "-",
             "SLA达标率", f"{report.sla_compliance_rate}%" if report.sla_compliance_rate else "-"],
        ]
        summary_table = Table(summary_data, colWidths=[4.5 * cm, 4 * cm, 4.5 * cm, 4 * cm])
        summary_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 10 * mm))

        trend_chart = self._generate_trend_chart(report.trend_data or [])
        if trend_chart:
            story.append(Paragraph("二、近7天异常趋势", h2_style))
            story.append(Spacer(1, 5 * mm))
            img = Image(trend_chart, width=15 * cm, height=7 * cm)
            story.append(img)
            story.append(Spacer(1, 8 * mm))

        story.append(PageBreak())

        if report.top_anomaly_types:
            pie_chart = self._generate_pie_chart(report.top_anomaly_types, "异常类型分布")
            if pie_chart:
                story.append(Paragraph("三、异常类型分布", h2_style))
                story.append(Spacer(1, 5 * mm))
                img = Image(pie_chart, width=10 * cm, height=7 * cm)
                story.append(img)
                story.append(Spacer(1, 8 * mm))

        if report.system_summary:
            story.append(Paragraph("四、Top受影响系统", h2_style))
            story.append(Spacer(1, 5 * mm))

            sys_data = [["排名", "系统名称", "异常数", "严重异常", "解决率"]]
            for idx, s in enumerate(report.system_summary[:10], 1):
                sys_data.append([
                    str(idx), s["system"], str(s["total"]), str(s["critical"]), f"{s['resolution_rate']}%"
                ])

            sys_table = Table(sys_data, colWidths=[1.5 * cm, 5 * cm, 2.5 * cm, 2.5 * cm, 3 * cm])
            sys_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#27ae60")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(sys_table)
            story.append(Spacer(1, 8 * mm))

        if report.team_summary:
            story.append(Paragraph("五、团队处理情况", h2_style))
            story.append(Spacer(1, 5 * mm))

            team_data = [["团队", "工单数", "完成数", "升级数", "完成率"]]
            for t in report.team_summary[:8]:
                team_data.append([
                    t["team"], str(t["total"]), str(t["completed"]), str(t["escalated"]), f"{t['completion_rate']}%"
                ])

            team_table = Table(team_data, colWidths=[4 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm])
            team_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8e44ad")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(team_table)
            story.append(Spacer(1, 8 * mm))

        if report.recommendations:
            story.append(Paragraph("六、优化建议", h2_style))
            story.append(Spacer(1, 5 * mm))
            for idx, rec in enumerate(report.recommendations, 1):
                story.append(Paragraph(f"{idx}. {rec}", normal_style))
            story.append(Spacer(1, 5 * mm))

        if report.notable_events:
            story.append(PageBreak())
            story.append(Paragraph("七、重要异常事件", h2_style))
            story.append(Spacer(1, 5 * mm))
            event_data = [["异常编号", "标题", "系统", "等级", "状态", "影响分"]]
            for e in report.notable_events:
                event_data.append([
                    e["anomaly_code"],
                    Paragraph(e["title"][:50], normal_style),
                    e["system"],
                    e["severity"],
                    e["status"],
                    f"{e['impact_score']:.0f}",
                ])
            event_table = Table(event_data, colWidths=[3.5 * cm, 5 * cm, 2.5 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm])
            event_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e74c3c")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            story.append(event_table)

        def add_footer(canvas_obj, doc_obj):
            canvas_obj.saveState()
            canvas_obj.setFont(font_name, 8)
            canvas_obj.drawCentredString(A4[0] / 2, 1 * cm, f"运维异常检测系统 - 每日报告 - {date_str}")
            canvas_obj.drawRightString(A4[0] - 2 * cm, 1 * cm, f"第 {doc_obj.page} 页")
            canvas_obj.restoreState()

        doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)

        async with async_session_maker() as db:
            db_report = await db.execute(
                select(DailyReport).where(DailyReport.id == report.id)
            )
            r = db_report.scalar_one_or_none()
            if r:
                r.pdf_file_path = filepath
                await db.commit()

        logger.info(f"PDF report generated: {filepath}")
        return filepath

    def _generate_trend_chart(self, trend_data: List[Dict]) -> Optional[str]:
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            import io

            fig, ax = plt.subplots(figsize=(8, 4))
            dates = [d["date"] for d in trend_data]
            counts = [d["anomaly_count"] for d in trend_data]

            ax.plot(dates, counts, marker="o", linewidth=2, color="#3498db")
            ax.fill_between(dates, counts, alpha=0.3, color="#3498db")

            for i, (d, c) in enumerate(zip(dates, counts)):
                ax.annotate(str(c), (d, c), textcoords="offset points", xytext=(0, 10), ha="center", fontsize=9)

            ax.set_xlabel("日期", fontsize=11)
            ax.set_ylabel("异常数量", fontsize=11)
            ax.set_title("近7天异常趋势", fontsize=13, fontweight="bold")
            ax.grid(True, alpha=0.3)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            plt.tight_layout()

            filename = f"trend_{uuid.uuid4().hex[:8]}.png"
            filepath = os.path.join(self._export_dir, filename)
            fig.savefig(filepath, dpi=150, bbox_inches="tight")
            plt.close(fig)
            return filepath
        except Exception as e:
            logger.error(f"Failed to generate trend chart: {e}")
            return None

    def _generate_pie_chart(self, data: List[Dict], title: str) -> Optional[str]:
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt

            labels = [d.get("type", d.get("system", str(i))) for i, d in enumerate(data[:8])]
            sizes = [d.get("count", 0) for d in data[:8]]
            total = sum(sizes)
            if total == 0:
                return None

            colors = ["#3498db", "#e74c3c", "#2ecc71", "#f39c12", "#9b59b6",
                      "#1abc9c", "#d35400", "#34495e"]

            fig, ax = plt.subplots(figsize=(6, 6))
            wedges, texts, autotexts = ax.pie(
                sizes, labels=labels, autopct="%1.1f%%", colors=colors[:len(sizes)],
                startangle=90, pctdistance=0.85
            )

            for text in texts:
                text.set_fontsize(9)
            for autotext in autotexts:
                autotext.set_fontsize(9)
                autotext.set_color("white")
                autotext.set_fontweight("bold")

            centre_circle = plt.Circle((0, 0), 0.70, fc="white")
            ax.add_artist(centre_circle)
            ax.set_title(title, fontsize=13, fontweight="bold")
            ax.axis("equal")
            plt.tight_layout()

            filename = f"pie_{uuid.uuid4().hex[:8]}.png"
            filepath = os.path.join(self._export_dir, filename)
            fig.savefig(filepath, dpi=150, bbox_inches="tight")
            plt.close(fig)
            return filepath
        except Exception as e:
            logger.error(f"Failed to generate pie chart: {e}")
            return None

    async def export_report_excel(self, report_id: str) -> Optional[str]:
        async with async_session_maker() as db:
            result = await db.execute(
                select(DailyReport).where(DailyReport.id == report_id)
            )
            report = result.scalar_one_or_none()
            if report is None:
                return None

        return self._generate_excel_report(report)

    def _generate_excel_report(self, report: DailyReport) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import LineChart, PieChart, Reference, BarChart

        date_str = report.report_date.strftime("%Y-%m-%d")
        filename = f"daily_report_{date_str}_{uuid.uuid4().hex[:8]}.xlsx"
        filepath = os.path.join(self._export_dir, filename)

        wb = Workbook()

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        ws1 = wb.active
        ws1.title = "概览"

        ws1["A1"] = f"运维异常检测系统 - 每日汇总报告 ({date_str})"
        ws1["A1"].font = Font(bold=True, size=14, color="1A5490")
        ws1.merge_cells("A1:F1")

        ws1.append([])
        summary_headers = ["指标", "数值", "指标", "数值", "指标", "数值"]
        ws1.append(summary_headers)
        for cell in ws1[3]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        summary_rows = [
            ["总异常数", report.total_anomalies, "总工单数", report.total_work_orders, "升级次数", report.escalated_count],
            ["严重异常", report.critical_anomalies, "高危异常", report.high_anomalies, "中等异常", report.medium_anomalies],
            ["已解决", report.resolved_anomalies, "未解决", report.unresolved_anomalies, "新增", report.new_anomalies],
            ["自动修复", report.auto_resolved, "手动修复", report.manual_resolved, "预案修复", report.playbook_resolved],
            ["平均修复(分)", str(report.avg_resolution_minutes or "-"),
             "P95修复(分)", str(report.p95_resolution_minutes or "-"),
             "重复率(%)", str(report.repeat_rate or "-")],
            ["SLA达标(%)", str(report.sla_compliance_rate or "-"),
             "SLA违规数", report.sla_breach_count,
             "完成工单", report.completed_work_orders],
        ]
        for row in summary_rows:
            ws1.append(row)
            for cell in ws1[ws1.max_row]:
                cell.alignment = center_align
                cell.border = thin_border

        for col in ["A", "B", "C", "D", "E", "F"]:
            ws1.column_dimensions[col].width = 18

        if report.trend_data:
            ws2 = wb.create_sheet("趋势数据")
            ws2.append(["日期", "异常数量"])
            for cell in ws2[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            for d in report.trend_data:
                ws2.append([d["date"], d["anomaly_count"]])
                for cell in ws2[ws2.max_row]:
                    cell.alignment = center_align
                    cell.border = thin_border

            chart = LineChart()
            chart.title = "近7天异常趋势"
            chart.y_axis.title = "异常数量"
            chart.x_axis.title = "日期"
            data = Reference(ws2, min_col=2, min_row=1, max_row=len(report.trend_data) + 1)
            cats = Reference(ws2, min_col=1, min_row=2, max_row=len(report.trend_data) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.style = 10
            ws2.add_chart(chart, "D2")

        if report.top_anomaly_types:
            ws3 = wb.create_sheet("异常类型")
            ws3.append(["异常类型", "数量"])
            for cell in ws3[1]:
                cell.font = header_font
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                cell.alignment = center_align
                cell.border = thin_border
            for t in report.top_anomaly_types:
                ws3.append([t.get("type", t.get("category", "")), t["count"]])
                for cell in ws3[ws3.max_row]:
                    cell.alignment = center_align
                    cell.border = thin_border

        if report.system_summary:
            ws4 = wb.create_sheet("系统统计")
            ws4.append(["系统名称", "总异常", "严重", "已解决", "解决率(%)"])
            for cell in ws4[1]:
                cell.font = header_font
                cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                cell.alignment = center_align
                cell.border = thin_border
            for s in report.system_summary:
                ws4.append([s["system"], s["total"], s["critical"], s["resolved"], s["resolution_rate"]])
                for cell in ws4[ws4.max_row]:
                    cell.alignment = center_align
                    cell.border = thin_border

        if report.team_summary:
            ws5 = wb.create_sheet("团队统计")
            ws5.append(["团队", "工单数", "完成数", "升级数", "完成率(%)"])
            for cell in ws5[1]:
                cell.font = header_font
                cell.fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
                cell.alignment = center_align
                cell.border = thin_border
            for t in report.team_summary:
                ws5.append([t["team"], t["total"], t["completed"], t["escalated"], t["completion_rate"]])
                for cell in ws5[ws5.max_row]:
                    cell.alignment = center_align
                    cell.border = thin_border

        if report.anomaly_hourly_distribution:
            ws6 = wb.create_sheet("时段分布")
            ws6.append(["小时", "异常数"])
            for cell in ws6[1]:
                cell.font = header_font
                cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                cell.alignment = center_align
                cell.border = thin_border
            for hour in range(24):
                ws6.append([f"{hour:02d}:00", report.anomaly_hourly_distribution.get(str(hour), 0)])
                for cell in ws6[ws6.max_row]:
                    cell.alignment = center_align
                    cell.border = thin_border

            bar = BarChart()
            bar.title = "异常24小时时段分布"
            bar.y_axis.title = "异常数量"
            bar.x_axis.title = "时段"
            data = Reference(ws6, min_col=2, min_row=1, max_row=25)
            cats = Reference(ws6, min_col=1, min_row=2, max_row=25)
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            bar.style = 11
            ws6.add_chart(bar, "D2")

        if report.notable_events:
            ws7 = wb.create_sheet("重要事件")
            ws7.append(["异常编号", "标题", "系统", "严重等级", "状态", "影响分", "检测时间"])
            for cell in ws7[1]:
                cell.font = header_font
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                cell.alignment = center_align
                cell.border = thin_border
            for e in report.notable_events:
                ws7.append([
                    e["anomaly_code"], e["title"], e["system"],
                    e["severity"], e["status"], e["impact_score"], e["detected_time"]
                ])
                for cell in ws7[ws7.max_row]:
                    cell.alignment = center_align
                    cell.border = thin_border

        if report.recommendations:
            ws8 = wb.create_sheet("优化建议")
            ws8.append(["序号", "建议内容"])
            for cell in ws8[1]:
                cell.font = header_font
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                cell.alignment = center_align
                cell.border = thin_border
            for idx, rec in enumerate(report.recommendations, 1):
                ws8.append([idx, rec])
                for cell in ws8[ws8.max_row]:
                    cell.border = thin_border
            ws8.column_dimensions["B"].width = 80

        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = True

        wb.save(filepath)

        logger.info(f"Excel report generated: {filepath}")
        return filepath


report_service = ReportService()
